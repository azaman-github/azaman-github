'''
The module is used to create an excel workbook with multiple worksheets from
a number of csv files.
'''
#
# Name     :csv_to_excel.py
# Overview :The module is used to create an excel workbook with multiple worksheets from
#           a number of csv files.

# Output   : Excel workbook in the directory of json files
# Notes    :
#
# History
# Date       Author      Description
#
# 06/05/2022 Arif Zaman  Initial creation
#
import logging
import sys
import os
import traceback
from openpyxl import Workbook
#
# globals
#
#global dir_path
#global wb
#
# module constant
#
#EXCEL_WB_NAME = f"{dir_path}/emis_patient_data.xlsx"
LOGIN = os.getlogin()
LOGFILE = "./log/elv.log"
MODULE_NAME = sys.argv[0]
#
# Name     :intialise_workbook
# Overview :The function is used to initialise a workbook using global variable, wb.
# Notes    :
#
def intialise_workbook ():
    '''
    The function is used to initialise a workbook using global variable, wb.
    '''
    global wb
    wb = Workbook()


#
# Name     :save_workbook
# Overview :The function is used to save the workbook in the current directory.
# Notes    :
#
def save_workbook () :
    '''
    The function is used to save the workbook in the current directory.
    '''
    global dir_path

    wb_name=f"{dir_path}/emis_patient_data.xlsx"
    wb.save(filename=wb_name )


#
# Name     :create_worksheet
# Overview :The function is used to create a new worksheet from a
#           given csv file
# Notes    :
#
def create_worksheet (csv_file) :
    '''
    The function is used to create a new worksheet from a
    given csv file
    '''
    #
    # define worksheet name
    #
    ws_name=os.path.basename(csv_file).split(".")[0]
    #print(f"ws_name={ws_name}")
    ws = wb.create_sheet(ws_name)
    next_row=1
    next_col=1
    with open (csv_file, "r") as fh:
        for rec in fh :
            lov=rec.split(",")
            for val in lov :
                cell = ws.cell(row = next_row, column = next_col)
                cell.value = val
                next_col += 1
            #
            next_col = 1
            next_row += 1


#
# Name     :main
# Overview :The entry function.
# Notes    :
#
def main () :
    '''
    The entry function.
    '''
    try:
        global dir_path
        #
        logging.root.name = LOGIN
        logging.basicConfig(
            level=logging.INFO, filename=LOGFILE, filemode="a",
            format="%(asctime)s %(name)s - %(levelname)s - %(message)s")

        logging.info(f"Module-{MODULE_NAME} Execution Starting")
        #
        # check for argument
        #
        if len(sys.argv)  != 2  :
            raise Exception("Usage:create_excel_spreadsheets_from_csv_files.py <dir path>")

        dir_path=sys.argv[1]
        #
        # validate directory path
        #
        if not os.path.exists(dir_path):
            raise Exception("ERROR:Invalid directory path")
        #
        # get all the csv file names
        # lof=list of files
        #
        lof=[fname for fname in  os.listdir(dir_path) if "csv" in  fname.split(".")]
        if len (lof) == 0 :
            raise Exception(f"ERROR:{dir_path} does not contain any csv files")

        intialise_workbook()

        for fname in lof :
            file_path=f"{dir_path}/{fname}"
            create_worksheet(file_path)

        save_workbook()

        logging.info(f"Module-{MODULE_NAME} Execution Completed")

    except Exception as e :
        logging.exception("Exception:")
        traceback.print_exc()

main ()
